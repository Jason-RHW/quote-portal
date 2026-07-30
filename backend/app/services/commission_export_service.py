"""
SDR Commission Dashboard — Excel export.

Builds the reviewed multi-sheet workbook (Summary, Sample Detail, Quote
Detail, hidden SPIFF Components, SPIFF, Deal Commission Detail) with every
number that can be derived from other cells written as a real Excel
formula rather than a pre-computed value — see build_commission_workbook's
docstring for the full sheet-by-sheet breakdown.
"""
import io
import re
from collections import defaultdict
from datetime import date as date_cls
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula
from sqlalchemy.orm import Session

from app.models.db_models import Quote
from app.services import spiff_service

SAMPLE_BASE_RATE = 1.0
QUOTE_BASE_RATE = 3.0
MEETING_BASE_RATE = 3.0

HEADER_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
GROUP_FONT = Font(bold=True, size=12, color="1E3A5F")
SUBTOTAL_FILL = PatternFill(start_color="EAF0F6", end_color="EAF0F6", fill_type="solid")
SUBTOTAL_FONT = Font(bold=True)
EMPTY_FONT = Font(italic=True, color="6B7280")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY_FMT = '"$"#,##0.00'
PERCENT_FMT = '0.00"%"'
DATE_FMT = "yyyy-mm-dd"

SAMPLE_RATE_REF = "Summary!$M$1"
QUOTE_RATE_REF = "Summary!$M$2"
MEETING_RATE_REF = "Summary!$M$3"
COMPONENTS_SHEET = "SPIFF Components"


def _style_header_row(ws, row_idx, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def _visible_text(value, fmt=None):
    if value is None or value == "":
        return ""
    if fmt == MONEY_FMT:
        try:
            return f"${float(value):,.2f}"
        except (TypeError, ValueError):
            return str(value)
    if fmt == PERCENT_FMT:
        try:
            return f"{float(value):.2f}%"
        except (TypeError, ValueError):
            return str(value)
    return str(value)


def _set_col_width(widths, col, text):
    widths[col] = max(widths.get(col, 0), len(text))


def _apply_col_widths(ws, widths, padding=2):
    for col, max_len in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = max_len + padding


_MONTH_NAMES = (
    r"(?:January|February|March|April|May|June|July|August|September|October|November|December"
    r"|Jan|Feb|Mar|Apr|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec)"
)
_MONTH_DAY_RANGE = (
    rf"{_MONTH_NAMES}\.?\s+\d{{1,2}}(?:st|nd|rd|th)?"
    rf"(?:\s*[-–—]\s*\d{{1,2}}(?:st|nd|rd|th)?)?(?:,?\s*\d{{4}})?"
)
_ISO_DATE = r"\d{4}-\d{2}-\d{2}"
_SLASH_DATE = r"\d{1,2}/\d{1,2}(?:/\d{2,4})?"
_DATE_TOKEN = rf"(?:{_MONTH_DAY_RANGE}|{_ISO_DATE}|{_SLASH_DATE})"
_PREP = r"(?:\b(?:on|for|during|in|of|at)\b\s*)?"
_DATE_RANGE_RE = re.compile(
    rf"{_PREP}\(?\s*{_DATE_TOKEN}\s*(?:[-–—]|to|through)\s*{_DATE_TOKEN}\s*\)?|{_PREP}\(?\s*{_DATE_TOKEN}\s*\)?",
    re.IGNORECASE,
)


def _strip_dates_from_name(name: Optional[str]) -> str:
    """Defensive cleanup for any campaign name that embeds a date/date
    range — the export always shows a single, consistently-formatted date
    range in its own column, so any date text baked into the name itself
    (from an AI-generated rule, or a manually-typed one) would be
    redundant or could disagree with the real start/end dates."""
    if not name:
        return name or ""
    cleaned = _DATE_RANGE_RE.sub("", name)
    cleaned = re.sub(r"\(\s*\)", "", cleaned)
    cleaned = re.sub(r"\s{2,}", " ", cleaned).strip(" -–—:,")
    return cleaned or name


def _parse_date(value):
    if not value:
        return None
    try:
        return date_cls.fromisoformat(value)
    except (TypeError, ValueError):
        return None


def _add_plain_list_sheet(wb, title, ordered_results, item_key, count_col=1, extra_field=None):
    """Just the raw records — Date, Business Name, optionally a third field
    (e.g. Quote Value), and a hidden Record ID column. This is the single
    source of truth the SPIFF sheet's Date/Business Name formulas look up
    by Record ID — editing a date/name here propagates everywhere. Returns
    count_ranges (for Summary's COUNTA formulas) and the ID column letter."""
    ws = wb.create_sheet(title)
    row = 1
    widths = defaultdict(int)
    count_ranges = {}
    header_labels = ["Date", "Business Name"]
    if extra_field:
        header_labels = header_labels + [extra_field[0]]
    header_labels = header_labels + ["Record ID"]
    id_col = len(header_labels)

    for r in ordered_results:
        sdr = r["sdr_name"]
        items = r.get(item_key) or []

        ws.cell(row=row, column=1, value=sdr).font = GROUP_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(header_labels))
        _set_col_width(widths, 1, sdr)
        row += 1

        if not items:
            ws.cell(row=row, column=1, value="No records this month").font = EMPTY_FONT
            _set_col_width(widths, 1, "No records this month")
            for col in range(1, len(header_labels) + 1):
                ws.cell(row=row, column=col).border = BORDER
            count_ranges[sdr] = None
            row += 2
            continue

        for col, label in enumerate(header_labels, start=1):
            ws.cell(row=row, column=col, value=label)
            _set_col_width(widths, col, label)
        _style_header_row(ws, row, len(header_labels))
        row += 1

        data_start_row = row
        for item in sorted(items, key=lambda x: x.get("date") or ""):
            values = [item.get("date"), item.get("business_name")]
            if extra_field:
                values.append(extra_field[1](item))
            values.append(item.get("id"))
            for col, val in enumerate(values, start=1):
                is_money = extra_field and col == 3
                fmt = MONEY_FMT if is_money else None
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = BORDER
                if fmt:
                    cell.number_format = fmt
                _set_col_width(widths, col, _visible_text(val, fmt))
            row += 1
        data_end_row = row - 1
        count_col_letter = get_column_letter(count_col)
        count_ranges[sdr] = f"'{title}'!{count_col_letter}{data_start_row}:{count_col_letter}{data_end_row}"

        row += 1  # spacer

    _apply_col_widths(ws, widths)
    ws.column_dimensions[get_column_letter(id_col)].hidden = True
    return count_ranges, get_column_letter(id_col)


def _add_grouped_sheet(wb, title, ordered_results, item_key, header_labels, row_fn, money_cols,
                        subtotal_col=None, percent_cols=(), formula_col=None, formula_fn=None):
    """Generic grouped-by-SDR sheet, used for Deal Commission Detail."""
    ws = wb.create_sheet(title)
    row = 1
    widths = defaultdict(int)
    subtotal_refs = {}

    def fmt_for(col):
        if col in money_cols:
            return MONEY_FMT
        if col in percent_cols:
            return PERCENT_FMT
        return None

    for r in ordered_results:
        sdr = r["sdr_name"]
        items = r.get(item_key) or []

        ws.cell(row=row, column=1, value=sdr).font = GROUP_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(header_labels))
        _set_col_width(widths, 1, sdr)
        row += 1

        if not items:
            ws.cell(row=row, column=1, value="No records this month").font = EMPTY_FONT
            _set_col_width(widths, 1, "No records this month")
            zero_col = subtotal_col or len(header_labels)
            zero_cell = ws.cell(row=row, column=zero_col, value=0)
            zero_cell.number_format = MONEY_FMT
            zero_cell.font = EMPTY_FONT
            _set_col_width(widths, zero_col, _visible_text(0, MONEY_FMT))
            for col in range(1, len(header_labels) + 1):
                ws.cell(row=row, column=col).border = BORDER
            col_letter = get_column_letter(zero_col)
            subtotal_refs[sdr] = f"'{title}'!{col_letter}{row}"
            row += 2
            continue

        for col, label in enumerate(header_labels, start=1):
            ws.cell(row=row, column=col, value=label)
            _set_col_width(widths, col, label)
        _style_header_row(ws, row, len(header_labels))
        row += 1

        data_start_row = row
        for item in sorted(items, key=lambda x: x.get("date") or ""):
            values = row_fn(item)
            for col, val in enumerate(values, start=1):
                fmt = fmt_for(col)
                if formula_col and col == formula_col and formula_fn:
                    cell = ws.cell(row=row, column=col, value=formula_fn(row))
                else:
                    cell = ws.cell(row=row, column=col, value=val)
                cell.border = BORDER
                if fmt:
                    cell.number_format = fmt
                _set_col_width(widths, col, _visible_text(val, fmt))
            row += 1
        data_end_row = row - 1

        if subtotal_col:
            label = f"{sdr} Subtotal"
            ws.cell(row=row, column=subtotal_col - 1, value=label).font = SUBTOTAL_FONT
            _set_col_width(widths, subtotal_col - 1, label)
            col_letter = get_column_letter(subtotal_col)
            numeric_total = sum((row_fn(item)[subtotal_col - 1] or 0) for item in items)
            subtotal_cell = ws.cell(row=row, column=subtotal_col, value=f"=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})")
            subtotal_cell.font = SUBTOTAL_FONT
            subtotal_cell.number_format = MONEY_FMT
            _set_col_width(widths, subtotal_col, _visible_text(numeric_total, MONEY_FMT))
            for col in range(1, len(header_labels) + 1):
                ws.cell(row=row, column=col).fill = SUBTOTAL_FILL
                ws.cell(row=row, column=col).border = BORDER
            subtotal_refs[sdr] = f"'{title}'!{col_letter}{row}"
            row += 1

        row += 1

    _apply_col_widths(ws, widths)
    return subtotal_refs


def _add_spiff_components_sheet(wb, ordered_results):
    """Hidden row-level source data backing the combined SPIFF sheet: every
    individual sample/quote/meeting SPIFF delta and every overall bonus
    instance, one row each. Start/End Date are real date objects (not text)
    so the array-formula MIN/MAX lookups on the visible sheet work — Excel's
    date-comparison functions ignore text values. Amount is a formula
    (Final Amount - base rate) for sample/quote/meeting rows; a literal for
    bonus rows (there's no base rate to subtract from a team/threshold
    bonus). Quote-linked meetings never appear here — they have no base
    rate to override, so they never carry spiff_campaigns."""
    ws = wb.create_sheet(COMPONENTS_SHEET)
    headers = ["SDR Name", "Campaign", "Start Date", "End Date", "Final Amount", "Amount", "Source"]
    for col, label in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=label)
    _style_header_row(ws, 1, len(headers))

    row = 2
    for r in ordered_results:
        sdr = r["sdr_name"]

        for record in r.get("samples") or []:
            for campaign in record.get("spiff_campaigns") or []:
                name = _strip_dates_from_name(campaign.get("name") or "SPIFF")
                start = _parse_date(campaign.get("start_date"))
                end = _parse_date(campaign.get("end_date"))
                final_amount = campaign.get("rate")
                ws.cell(row=row, column=1, value=sdr)
                ws.cell(row=row, column=2, value=name)
                c3 = ws.cell(row=row, column=3, value=start)
                if start:
                    c3.number_format = DATE_FMT
                c4 = ws.cell(row=row, column=4, value=end)
                if end:
                    c4.number_format = DATE_FMT
                c5 = ws.cell(row=row, column=5, value=final_amount)
                c5.number_format = MONEY_FMT
                c6 = ws.cell(row=row, column=6, value=f"=E{row}-{SAMPLE_RATE_REF}")
                c6.number_format = MONEY_FMT
                ws.cell(row=row, column=7, value="Sample")
                row += 1

        for record in r.get("quotes") or []:
            for campaign in record.get("spiff_campaigns") or []:
                name = _strip_dates_from_name(campaign.get("name") or "SPIFF")
                start = _parse_date(campaign.get("start_date"))
                end = _parse_date(campaign.get("end_date"))
                final_amount = campaign.get("rate")
                ws.cell(row=row, column=1, value=sdr)
                ws.cell(row=row, column=2, value=name)
                c3 = ws.cell(row=row, column=3, value=start)
                if start:
                    c3.number_format = DATE_FMT
                c4 = ws.cell(row=row, column=4, value=end)
                if end:
                    c4.number_format = DATE_FMT
                c5 = ws.cell(row=row, column=5, value=final_amount)
                c5.number_format = MONEY_FMT
                c6 = ws.cell(row=row, column=6, value=f"=E{row}-{QUOTE_RATE_REF}")
                c6.number_format = MONEY_FMT
                ws.cell(row=row, column=7, value="Quote")
                row += 1

        for record in r.get("meetings") or []:
            for campaign in record.get("spiff_campaigns") or []:
                name = _strip_dates_from_name(campaign.get("name") or "SPIFF")
                start = _parse_date(campaign.get("start_date"))
                end = _parse_date(campaign.get("end_date"))
                final_amount = campaign.get("rate")
                ws.cell(row=row, column=1, value=sdr)
                ws.cell(row=row, column=2, value=name)
                c3 = ws.cell(row=row, column=3, value=start)
                if start:
                    c3.number_format = DATE_FMT
                c4 = ws.cell(row=row, column=4, value=end)
                if end:
                    c4.number_format = DATE_FMT
                c5 = ws.cell(row=row, column=5, value=final_amount)
                c5.number_format = MONEY_FMT
                c6 = ws.cell(row=row, column=6, value=f"=E{row}-{MEETING_RATE_REF}")
                c6.number_format = MONEY_FMT
                ws.cell(row=row, column=7, value="Meeting")
                row += 1

        for bonus in r.get("spiff_bonus_details") or []:
            name = _strip_dates_from_name(bonus.get("name") or "SPIFF")
            start = _parse_date(bonus.get("start_date"))
            end = _parse_date(bonus.get("end_date"))
            ws.cell(row=row, column=1, value=sdr)
            ws.cell(row=row, column=2, value=name)
            c3 = ws.cell(row=row, column=3, value=start)
            if start:
                c3.number_format = DATE_FMT
            c4 = ws.cell(row=row, column=4, value=end)
            if end:
                c4.number_format = DATE_FMT
            ws.cell(row=row, column=6, value=bonus.get("amount"))
            ws.cell(row=row, column=6).number_format = MONEY_FMT
            ws.cell(row=row, column=7, value="Overall Bonus")
            row += 1

    last_row = row - 1
    for col, width in zip(range(1, 8), [26, 30, 12, 12, 12, 12, 14]):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.sheet_state = "hidden"
    return last_row


def _add_combined_spiff_sheet(wb, ordered_results, components_last_row):
    """One row per (SDR, campaign) — aggregated across samples, quotes, and
    overall bonuses that share that campaign name. Campaign/Start
    Date/End Date/SPIFF $ are all formulas against the hidden SPIFF
    Components sheet: SUMIFS for the dollar total (widely supported since
    Excel 2007), and MIN(IF())/MAX(IF()) array formulas for the date range
    (MINIFS/MAXIFS aren't recognized by every Excel version — this classic
    equivalent has worked since Excel 97)."""
    ws = wb.create_sheet("SPIFF")
    row = 1
    widths = defaultdict(int)
    subtotal_refs = {}
    header_labels = ["Campaign", "Start Date", "End Date", "SPIFF $"]
    CAMPAIGN_COL, START_COL, END_COL, AMOUNT_COL, SDR_HELPER_COL = 1, 2, 3, 4, 5

    def comp_range(col):
        return f"'{COMPONENTS_SHEET}'!${col}$2:${col}${components_last_row}"

    SDR_RANGE = comp_range("A")
    CAMPAIGN_RANGE = comp_range("B")
    START_RANGE = comp_range("C")
    END_RANGE = comp_range("D")
    AMOUNT_RANGE = comp_range("F")

    for r in ordered_results:
        sdr = r["sdr_name"]

        campaign_totals = defaultdict(float)
        for record in r.get("samples") or []:
            for campaign in record.get("spiff_campaigns") or []:
                name = _strip_dates_from_name(campaign.get("name") or "SPIFF")
                campaign_totals[name] += float(campaign.get("rate") or 0) - SAMPLE_BASE_RATE
        for record in r.get("quotes") or []:
            for campaign in record.get("spiff_campaigns") or []:
                name = _strip_dates_from_name(campaign.get("name") or "SPIFF")
                campaign_totals[name] += float(campaign.get("rate") or 0) - QUOTE_BASE_RATE
        for bonus in r.get("spiff_bonus_details") or []:
            name = _strip_dates_from_name(bonus.get("name") or "SPIFF")
            campaign_totals[name] += float(bonus.get("amount") or 0)
        campaign_names_sorted = sorted(campaign_totals, key=lambda n: -campaign_totals[n])

        ws.cell(row=row, column=1, value=sdr).font = GROUP_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(header_labels))
        _set_col_width(widths, 1, sdr)
        row += 1

        if not campaign_names_sorted:
            ws.cell(row=row, column=1, value="No records this month").font = EMPTY_FONT
            _set_col_width(widths, 1, "No records this month")
            zero_cell = ws.cell(row=row, column=AMOUNT_COL, value=0)
            zero_cell.number_format = MONEY_FMT
            zero_cell.font = EMPTY_FONT
            _set_col_width(widths, AMOUNT_COL, _visible_text(0, MONEY_FMT))
            for col in range(1, len(header_labels) + 1):
                ws.cell(row=row, column=col).border = BORDER
            col_letter = get_column_letter(AMOUNT_COL)
            subtotal_refs[sdr] = f"'SPIFF'!{col_letter}{row}"
            row += 2
            continue

        for col, label in enumerate(header_labels, start=1):
            ws.cell(row=row, column=col, value=label)
            _set_col_width(widths, col, label)
        _style_header_row(ws, row, len(header_labels))
        row += 1

        data_start_row = row
        for name in campaign_names_sorted:
            ws.cell(row=row, column=CAMPAIGN_COL, value=name).border = BORDER
            _set_col_width(widths, CAMPAIGN_COL, name)

            sdr_helper = ws.cell(row=row, column=SDR_HELPER_COL, value=sdr)
            sdr_helper.border = BORDER
            sdr_cell_ref = f"$E{row}"

            match_mask = f"(({SDR_RANGE}={sdr_cell_ref})*({CAMPAIGN_RANGE}=$A{row}))"
            start_ref = f"{get_column_letter(START_COL)}{row}"
            end_ref = f"{get_column_letter(END_COL)}{row}"
            start_formula = ArrayFormula(start_ref, f"=MIN(IF({match_mask},{START_RANGE}))")
            end_formula = ArrayFormula(end_ref, f"=MAX(IF({match_mask},{END_RANGE}))")
            amount_formula = f"=SUMIFS({AMOUNT_RANGE},{SDR_RANGE},{sdr_cell_ref},{CAMPAIGN_RANGE},$A{row})"

            start_cell = ws.cell(row=row, column=START_COL, value=start_formula)
            start_cell.number_format = DATE_FMT
            start_cell.border = BORDER
            end_cell = ws.cell(row=row, column=END_COL, value=end_formula)
            end_cell.number_format = DATE_FMT
            end_cell.border = BORDER
            amount_cell = ws.cell(row=row, column=AMOUNT_COL, value=amount_formula)
            amount_cell.number_format = MONEY_FMT
            amount_cell.border = BORDER

            _set_col_width(widths, START_COL, "2026-07-13")
            _set_col_width(widths, END_COL, "2026-07-13")
            _set_col_width(widths, AMOUNT_COL, _visible_text(campaign_totals[name], MONEY_FMT))
            row += 1
        data_end_row = row - 1

        label = f"{sdr} Additional Prize"
        ws.cell(row=row, column=AMOUNT_COL - 1, value=label).font = SUBTOTAL_FONT
        _set_col_width(widths, AMOUNT_COL - 1, label)
        col_letter = get_column_letter(AMOUNT_COL)
        subtotal_cell = ws.cell(row=row, column=AMOUNT_COL, value=f"=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})")
        subtotal_cell.font = SUBTOTAL_FONT
        subtotal_cell.number_format = MONEY_FMT
        for col in range(1, len(header_labels) + 1):
            ws.cell(row=row, column=col).fill = SUBTOTAL_FILL
            ws.cell(row=row, column=col).border = BORDER
        subtotal_refs[sdr] = f"'SPIFF'!{col_letter}{row}"
        row += 1
        row += 1  # spacer

    _apply_col_widths(ws, widths)
    ws.column_dimensions[get_column_letter(SDR_HELPER_COL)].hidden = True
    return subtotal_refs


def _add_sick_days_sheet(wb, ordered_results):
    """Record-keeping only sheet — no dollar figures, so it doesn't reuse
    _add_grouped_sheet (whose empty-state always writes a $-formatted zero,
    which would be wrong here)."""
    ws = wb.create_sheet("Sick Days")
    header_labels = ["Start Date", "End Date", "Reason Note"]
    row = 1
    widths = defaultdict(int)

    for r in ordered_results:
        sdr = r["sdr_name"]
        rows = r.get("sick_days") or []

        ws.cell(row=row, column=1, value=sdr).font = GROUP_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(header_labels))
        _set_col_width(widths, 1, sdr)
        row += 1

        if not rows:
            ws.cell(row=row, column=1, value="No sick days this month").font = EMPTY_FONT
            _set_col_width(widths, 1, "No sick days this month")
            for col in range(1, len(header_labels) + 1):
                ws.cell(row=row, column=col).border = BORDER
            row += 2
            continue

        for col, label in enumerate(header_labels, start=1):
            ws.cell(row=row, column=col, value=label)
            _set_col_width(widths, col, label)
        _style_header_row(ws, row, len(header_labels))
        row += 1

        for item in sorted(rows, key=lambda x: x.get("start_date") or ""):
            values = [item.get("start_date"), item.get("end_date"), item.get("reason_note")]
            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = BORDER
                _set_col_width(widths, col, _visible_text(val))
            row += 1

        row += 1  # spacer

    _apply_col_widths(ws, widths)
    return None


def build_commission_workbook(db: Session, month: str) -> Workbook:
    """Sheet layout:
      - Summary (protected/locked): per-SDR totals. Sample $/Quote $ =
        count * base-rate (formula, base rates in labeled cells M1/M2/M3);
        Meeting $/SPIFF = cross-sheet reference to the Meeting Detail /
        combined SPIFF sheet's subtotal; Total Payout = same-row SUM;
        TOTAL row = SUM down each column.
      - Sample Detail / Quote Detail: plain record list (Date, Business
        Name, [Quote Value for quotes], hidden Record ID) — the single
        source of truth other sheets look up by ID.
      - SPIFF: one row per (SDR, campaign), aggregated across samples,
        quotes, meetings, and overall bonuses — Campaign, Start Date, End
        Date, SPIFF $ — all formulas against the hidden SPIFF Components
        sheet.
      - Deal Commission Detail: Commission $ = Deal Value * Commission % /
        100 (formula); each SDR's subtotal = SUM() over their own rows.
      - Meeting Detail: Date / Business Name / Source (Quote-Linked or
        Manual) / Amount — Amount is already the final, rule-aware number
        from the report, same as Sample/Quote Detail's plain amounts.
      - Sick Days: Start Date / End Date / Reason Note, no dollar figure —
        record-keeping only, not part of any payout math.
    """
    campaigns = spiff_service.list_campaigns(db, month)
    rules = [c["rule"] for c in campaigns]
    report = spiff_service.apply_rules_to_month(db, month, rules) if rules else spiff_service.monthly_dashboard(db, month)
    results = report["results"]

    quote_values = {str(q.id): float(q.quote_value or 0) for q in db.query(Quote).all()}

    wb = Workbook()
    wb.remove(wb.active)

    sample_counts, sample_id_col = _add_plain_list_sheet(wb, "Sample Detail", results, "samples")
    quote_counts, quote_id_col = _add_plain_list_sheet(
        wb, "Quote Detail", results, "quotes",
        extra_field=("Quote Value", lambda item: quote_values.get(item.get("id"), 0)),
    )

    components_last_row = _add_spiff_components_sheet(wb, results)
    spiff_subtotals = _add_combined_spiff_sheet(wb, results, components_last_row)

    deal_subtotals = _add_grouped_sheet(
        wb, "Deal Commission Detail", results, "deals",
        ["Date", "Business Name", "Deal Value", "Commission %", "Commission $"],
        lambda d: [d.get("date"), d.get("business_name"), d.get("deal_value"), d.get("commission_pct"), d.get("amount")],
        money_cols={3, 5}, subtotal_col=5, percent_cols={4},
        formula_col=5, formula_fn=lambda row: f"=C{row}*D{row}/100",
    )

    meeting_subtotals = _add_grouped_sheet(
        wb, "Meeting Detail", results, "meetings",
        ["Date", "Business Name", "Source", "Amount"],
        lambda m: [m.get("date"), m.get("business_name"), "Quote-Linked" if m.get("source_quote_id") else "Manual", m.get("amount")],
        money_cols={4}, subtotal_col=4,
    )

    _add_sick_days_sheet(wb, results)

    # ── Summary ──
    ws = wb.create_sheet("Summary")
    ws.append(["SDR Commission Summary — " + month])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    headers = ["SDR Name", "Samples", "Sample $", "Quotes", "Quote $", "Meeting $", "SPIFF", "Deal Commission $", "Total Payout"]
    ws.append(headers)
    _style_header_row(ws, 3, len(headers))
    money_col_idxs = {3, 5, 6, 7, 8, 9}
    widths = defaultdict(int)
    for col, label in enumerate(headers, start=1):
        _set_col_width(widths, col, label)

    ws.cell(row=1, column=12, value="Base rate per sample")
    rate_cell_1 = ws.cell(row=1, column=13, value=SAMPLE_BASE_RATE)
    rate_cell_1.number_format = MONEY_FMT
    ws.cell(row=2, column=12, value="Base rate per quote")
    rate_cell_2 = ws.cell(row=2, column=13, value=QUOTE_BASE_RATE)
    rate_cell_2.number_format = MONEY_FMT
    ws.cell(row=3, column=12, value="Base rate per meeting")
    rate_cell_3 = ws.cell(row=3, column=13, value=MEETING_BASE_RATE)
    rate_cell_3.number_format = MONEY_FMT
    _set_col_width(widths, 12, "Base rate per meeting")
    _set_col_width(widths, 13, _visible_text(QUOTE_BASE_RATE, MONEY_FMT))

    first_data_row = 4
    for r in results:
        sdr = r["sdr_name"]
        row_idx = ws.max_row + 1
        ws.cell(row=row_idx, column=1, value=sdr)

        sample_range = sample_counts.get(sdr)
        ws.cell(row=row_idx, column=2, value=f"=COUNTA({sample_range})" if sample_range else 0)
        quote_range = quote_counts.get(sdr)
        ws.cell(row=row_idx, column=4, value=f"=COUNTA({quote_range})" if quote_range else 0)

        ws.cell(row=row_idx, column=3, value=f"=B{row_idx}*$M$1")
        ws.cell(row=row_idx, column=5, value=f"=D{row_idx}*$M$2")
        ws.cell(row=row_idx, column=6, value=f"={meeting_subtotals[sdr]}")
        ws.cell(row=row_idx, column=7, value=f"={spiff_subtotals[sdr]}")
        ws.cell(row=row_idx, column=8, value=f"={deal_subtotals[sdr]}")
        ws.cell(row=row_idx, column=9, value=f"=C{row_idx}+E{row_idx}+F{row_idx}+G{row_idx}+H{row_idx}")

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = BORDER
            if col in money_col_idxs:
                cell.number_format = MONEY_FMT

        spiff_total = (r["sample_payout"] - r["eligible_sample_count"] * SAMPLE_BASE_RATE) \
            + (r["quote_payout"] - r["eligible_quote_count"] * QUOTE_BASE_RATE) + r["spiff_payout"]
        meeting_total = r.get("meeting_payout", 0)
        display_row = [
            sdr, r["eligible_sample_count"], r["eligible_sample_count"] * SAMPLE_BASE_RATE,
            r["eligible_quote_count"], r["eligible_quote_count"] * QUOTE_BASE_RATE,
            meeting_total, spiff_total,
            r.get("deal_payout", 0),
            r["eligible_sample_count"] * SAMPLE_BASE_RATE + r["eligible_quote_count"] * QUOTE_BASE_RATE
            + meeting_total + spiff_total + r.get("deal_payout", 0),
        ]
        for col, val in enumerate(display_row, start=1):
            fmt = MONEY_FMT if col in money_col_idxs else None
            _set_col_width(widths, col, _visible_text(val, fmt))

    last_data_row = ws.max_row
    total_row_idx = last_data_row + 1
    ws.cell(row=total_row_idx, column=1, value="TOTAL")
    for col in range(2, len(headers) + 1):
        col_letter = get_column_letter(col)
        ws.cell(row=total_row_idx, column=col, value=f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row_idx, column=col)
        cell.font = SUBTOTAL_FONT
        cell.fill = SUBTOTAL_FILL
        cell.border = BORDER
        if col in money_col_idxs:
            cell.number_format = MONEY_FMT

    ws.freeze_panes = "A4"
    _apply_col_widths(ws, widths)
    # Locked so it can't be accidentally edited/typed over — no password, so
    # anyone who genuinely needs to (Review > Unprotect Sheet) still can.
    ws.protection.sheet = True
    wb.move_sheet("Summary", offset=-(len(wb.sheetnames) - 1))

    return wb


def export_commission_excel(db: Session, month: str) -> bytes:
    wb = build_commission_workbook(db, month)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
