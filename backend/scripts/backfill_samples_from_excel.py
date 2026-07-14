"""
Backfill historical sample requests from the Sample Distribution Excel file.

Run from the backend directory:

    python scripts/backfill_samples_from_excel.py /Users/jason/Downloads/Sample_Distribution_MKT-6.xlsx --dry-run --limit 5
    python scripts/backfill_samples_from_excel.py /Users/jason/Downloads/Sample_Distribution_MKT-6.xlsx --limit 25

Optional API tests:

    # Runs OpenAI address verification on imported rows. Use a small --limit first.
    python scripts/backfill_samples_from_excel.py /Users/jason/Downloads/Sample_Distribution_MKT-6.xlsx --limit 5 --verify-addresses

    # After records with tracking are imported, also run HubSpot sync on eligible Sent/Delivered rows.
    python scripts/backfill_samples_from_excel.py /Users/jason/Downloads/Sample_Distribution_MKT-6.xlsx --limit 5 --sync-hubspot

The script is idempotent: it skips rows already present by
(contact_email, business_name, requested_date).

For historical HubSpot state:
- "Hubspot Note Processed" = tracking note was already synced and lifecycle
  moved to Sample Sent, so hubspot_sent_synced=True.
- "HubSpot Deliver Note Processed" = delivery note was already synced and
  lifecycle moved to Sample Delivered, so hubspot_delivered_synced=True.
"""
import argparse
import sys
import uuid
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.database import Base, SessionLocal, engine, is_sqlite
from app.models.db_models import (
    AddressVerificationStatus,
    Brand,
    FormField,
    SampleRequest,
    SampleRequestBrand,
    SampleRequestEvent,
    SampleRequestStatus,
    Sdr,
)
from app.services import sample_service


COLS = {
    "contact_name": "Full Name",
    "contact_email": "Email",
    "contact_phone": "Phone",
    "business_name": "Business Name",
    "address_line": "Address Line",
    "city": "City",
    "state": "State",
    "zip_code": "Zip Code",
    "size": "Size",
    "glove_type": "Glove Type",
    "color": "Color",
    "employee_count": "Number of Employees Using Gloves",
    "daily_changes": "Daily Glove Changes per Employee",
    "current_supplier": "Current Glove Brand or Supplier (if any)",
    "sample_note": "Type of Sample Notes",
    "product_sent": "Product Sent",
    "sample_sent": "Sample Sent",
    "status": "Status",
    "tracking_number": "Tracking ID",
    "delivery_status": "Delivered Date/Delivery Status",
    "note": "Note",
    "hubspot_note_processed": "Hubspot Note Processed",
    "sales_owner": "Sales Owner",
    "form_submitted_at": "Form Submiited At",
    "delivered_date": "Delivered Date",
    "hubspot_deliver_note_processed": "HubSpot Deliver Note Processed",
}

SDR_ALIASES = {
    "maria palmares": "Maria Gladys Palmares",
    "lorenzo bamiano": "Lhoreto Bamiano",
    "lhoreto bamiano": "Lhoreto Bamiano",
}


def clean(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.replace("\xa0", " ").strip()
        return value or None
    return str(value).strip() or None


def as_date(value: Any) -> Optional[date]:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def as_int_text(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return clean(value)


def norm(value: str) -> str:
    return " ".join(value.lower().split())


def is_processed(value: Any) -> bool:
    text = norm(clean(value) or "")
    if not text:
        return False
    if text in {"0", "false", "no", "n", "not processed", "unprocessed"}:
        return False
    return "processed" in text or text in {"1", "true", "yes", "y", "done", "complete", "completed"}


def status_from_excel(
    value: Any,
    tracking_number: Optional[str],
    delivered_date: Optional[date],
    delivery_status: Any = None,
    hubspot_sent_synced: bool = False,
    hubspot_delivered_synced: bool = False,
) -> SampleRequestStatus:
    text = (clean(value) or "").lower()
    delivery_text = (clean(delivery_status) or "").lower()
    if delivered_date or hubspot_delivered_synced or "deliver" in text or "deliver" in delivery_text:
        return SampleRequestStatus.delivered
    if tracking_number or hubspot_sent_synced or "sent" in text:
        return SampleRequestStatus.sent
    if "hold" in text:
        return SampleRequestStatus.on_hold
    if "reject" in text or "cancel" in text:
        return SampleRequestStatus.rejected
    return SampleRequestStatus.requested


def ensure_sdr(db, owner_name: Optional[str], create_missing: bool) -> Optional[str]:
    if not owner_name:
        return None
    canonical = SDR_ALIASES.get(norm(owner_name), owner_name)
    existing = db.query(Sdr).filter(Sdr.full_name == canonical).first()
    if existing:
        return existing.id
    if not create_missing:
        return None
    sdr = Sdr(id=str(uuid.uuid4()), full_name=canonical, active=False)
    db.add(sdr)
    db.flush()
    return sdr.id


def brand_ids_from_product(db, product_sent: Optional[str]) -> list[str]:
    if not product_sent:
        return []
    product = product_sent.lower()
    brands = db.query(Brand).all()
    ids = []
    for brand in brands:
        brand_key = brand.name.lower()
        if brand_key in product or brand_key.replace("flex", "flex") in product:
            ids.append(brand.id)
    return ids


def custom_fields_from(row: dict[str, Any]) -> dict[str, Any]:
    fields = {}
    mappings = {
        "glove_type": "glove_type",
        "size": "size",
        "color": "color",
        "employee_count": "employee_count",
        "daily_changes": "daily_changes",
        "current_supplier": "current_supplier",
    }
    for source, target in mappings.items():
        value = clean(row.get(COLS[source]))
        if value:
            if target in {"glove_type", "color"} and "," in value:
                fields[target] = [v.strip() for v in value.split(",") if v.strip()]
            else:
                fields[target] = value
    notes = [clean(row.get(COLS["sample_note"])), clean(row.get(COLS["note"]))]
    notes = [n for n in notes if n]
    if notes:
        fields["custom_requirement"] = " | ".join(notes)
    return fields


def already_exists(db, email: Optional[str], business_name: str, requested_date: date) -> bool:
    query = db.query(SampleRequest).filter(
        SampleRequest.business_name == business_name,
        SampleRequest.requested_date == requested_date,
    )
    if email:
        query = query.filter(SampleRequest.contact_email == email)
    return db.query(query.exists()).scalar()


def load_rows(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=False, data_only=True)
    ws = wb.active
    headers = [clean(cell.value) for cell in ws[1]]
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        if clean(row.get(COLS["business_name"])):
            rows.append(row)
    return rows


def insert_request(db, row: dict[str, Any], create_missing_sdrs: bool) -> Optional[SampleRequest]:
    business_name = clean(row.get(COLS["business_name"]))
    if not business_name:
        return None

    contact_email = clean(row.get(COLS["contact_email"]))
    requested_date = as_date(row.get(COLS["form_submitted_at"])) or date.today()
    if already_exists(db, contact_email, business_name, requested_date):
        return None

    tracking_number = as_int_text(row.get(COLS["tracking_number"]))
    delivered_date = as_date(row.get(COLS["delivered_date"])) or as_date(row.get(COLS["delivery_status"]))
    hubspot_sent_synced = is_processed(row.get(COLS["hubspot_note_processed"]))
    hubspot_delivered_synced = is_processed(row.get(COLS["hubspot_deliver_note_processed"]))
    if hubspot_delivered_synced:
        hubspot_sent_synced = True
    status = status_from_excel(
        row.get(COLS["status"]),
        tracking_number,
        delivered_date,
        row.get(COLS["delivery_status"]),
        hubspot_sent_synced,
        hubspot_delivered_synced,
    )
    sent_date = as_date(row.get(COLS["sample_sent"])) or (
        requested_date if status in {SampleRequestStatus.sent, SampleRequestStatus.delivered} else None
    )
    product_sent = clean(row.get(COLS["product_sent"])) or clean(row.get(COLS["sample_sent"]))
    brand_ids = brand_ids_from_product(db, product_sent)

    req = SampleRequest(
        id=str(uuid.uuid4()),
        sdr_id=ensure_sdr(db, clean(row.get(COLS["sales_owner"])), create_missing_sdrs),
        contact_name=clean(row.get(COLS["contact_name"])),
        contact_email=contact_email,
        contact_phone=as_int_text(row.get(COLS["contact_phone"])),
        business_name=business_name,
        address_line=clean(row.get(COLS["address_line"])),
        city=clean(row.get(COLS["city"])),
        state=clean(row.get(COLS["state"])),
        zip_code=as_int_text(row.get(COLS["zip_code"])),
        requested_date=requested_date,
        status=status,
        tracking_number=tracking_number,
        sent_date=sent_date,
        delivered_date=delivered_date,
        assignment_note=product_sent,
        custom_fields=custom_fields_from(row),
        address_verification_status=AddressVerificationStatus.unverified,
        hubspot_sent_synced=hubspot_sent_synced,
        hubspot_delivered_synced=hubspot_delivered_synced,
    )
    db.add(req)
    db.flush()
    for brand_id in brand_ids:
        db.add(SampleRequestBrand(id=str(uuid.uuid4()), sample_request_id=req.id, brand_id=brand_id))
    db.add(SampleRequestEvent(
        id=str(uuid.uuid4()),
        sample_request_id=req.id,
        from_status=None,
        to_status=status.value,
        changed_by="Excel backfill",
        note=(
            f"Imported from Sample Distribution Excel. Product Sent: {product_sent or '-'}; "
            f"HubSpot sent synced: {'yes' if hubspot_sent_synced else 'no'}; "
            f"HubSpot delivered synced: {'yes' if hubspot_delivered_synced else 'no'}"
        ),
    ))
    return req


def hubspot_sync_needed(req: SampleRequest) -> bool:
    if req.status == SampleRequestStatus.sent:
        return bool(req.tracking_number and not req.hubspot_sent_synced)
    if req.status == SampleRequestStatus.delivered:
        return not req.hubspot_delivered_synced
    return False


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx_path", type=Path)
    parser.add_argument("--limit", type=int, default=None, help="Import at most this many new records.")
    parser.add_argument("--dry-run", action="store_true", help="Preview without writing.")
    parser.add_argument("--verify-addresses", action="store_true", help="Run OpenAI address verification on imported rows.")
    parser.add_argument("--sync-hubspot", action="store_true", help="Run HubSpot sync on imported sent/delivered rows with tracking.")
    parser.add_argument("--no-create-missing-sdrs", action="store_true", help="Do not create historical SDR names if missing.")
    parser.add_argument("--create-tables", action="store_true", help="Create missing tables first. Defaults to on only for local SQLite.")
    args = parser.parse_args()

    if is_sqlite or args.create_tables:
        Base.metadata.create_all(bind=engine)
    rows = load_rows(args.xlsx_path)
    db = SessionLocal()
    imported: list[SampleRequest] = []

    try:
        for row in rows:
            if args.limit and len(imported) >= args.limit:
                break
            req = insert_request(db, row, create_missing_sdrs=not args.no_create_missing_sdrs)
            if req:
                imported.append(req)
                print(f"import: {req.business_name} | {req.contact_email or '-'} | {req.status.value} | {req.tracking_number or '-'}")

        if args.dry_run:
            db.rollback()
            print(f"DRY RUN: would import {len(imported)} new records.")
            return

        db.commit()
        print(f"Imported {len(imported)} new records.")

        if args.verify_addresses:
            print("Running address verification...")
            for req in imported:
                sample_service.rerun_address_verification(db, req.id)
            print(f"Address verification attempted for {len(imported)} records.")

        if args.sync_hubspot:
            eligible_ids = [
                req.id for req in imported
                if hubspot_sync_needed(req)
            ]
            print(f"Running HubSpot sync for {len(eligible_ids)} eligible records...")
            synced = sample_service.batch_hubspot_sync(db, eligible_ids)
            print(f"HubSpot synced {synced} records.")
    finally:
        db.close()


if __name__ == "__main__":
    main()
